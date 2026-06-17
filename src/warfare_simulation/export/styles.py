"""
Centralized style manager for spreadsheet generation.

Provides consistent fonts, fills, and alignments for all sheets,
matching the style of the original `campaign_engine_initialiser.py`.
"""

from openpyxl.styles import Alignment, Font, PatternFill


class StyleManager:
    """Centralized manager for spreadsheet cell styles."""

    @staticmethod
    def header_font() -> Font:
        """Returns the font for header cells."""
        return Font(bold=True, color="FFFFFF")

    @staticmethod
    def header_fill() -> PatternFill:
        """Returns the fill for header cells."""
        return PatternFill(start_color="333333", end_color="333333", fill_type="solid")

    @staticmethod
    def center_alignment() -> Alignment:
        """Returns a center-aligned alignment object."""
        return Alignment(horizontal="center", vertical="center")