"""Integration tests comparing modular export to the legacy monolith."""

from pathlib import Path

import openpyxl
import pytest

from campaign_engine_initialiser import create_campaign_workbook
from warfare_simulation.config.config import ConfigManager
from warfare_simulation.export.workbook_factory import WorkbookFactory
from warfare_simulation.persistence.campaign_bootstrap import CampaignBootstrap
from warfare_simulation.persistence.database import DatabaseManager

ROOT = Path(__file__).resolve().parents[1]
CONFIG_DIR = ROOT / "src" / "warfare_simulation" / "config" / "data"


@pytest.fixture
def legacy_workbook(tmp_path, monkeypatch):
    """Create the legacy workbook in an isolated directory and return it."""
    monkeypatch.chdir(tmp_path)
    workbook = create_campaign_workbook()
    if workbook is not None:
        return workbook
    return openpyxl.load_workbook(tmp_path / "Auster_Campaign_Engine.xlsx", data_only=False)


@pytest.fixture
def campaign_repos(tmp_path):
    """Create hydrated repositories from JSON -> SQLite runtime state."""
    db = DatabaseManager(str(tmp_path / "campaign.db"))
    config_mgr = ConfigManager(str(CONFIG_DIR))
    return CampaignBootstrap.initialize(config_mgr, db, force=True)


@pytest.fixture
def modular_workbook(campaign_repos):
    """Create a modular workbook from hydrated repositories."""
    factory = WorkbookFactory.from_campaign_repositories(campaign_repos)
    return factory.create_workbook()


def _non_empty_row_count(ws):
    return sum(1 for row in ws.iter_rows(values_only=True) if any(cell is not None for cell in row))


def _style_snapshot(cell):
    return {
        "font_bold": cell.font.bold,
        "font_color": cell.font.color.rgb if cell.font.color else None,
        "fill_start": cell.fill.start_color.rgb,
        "fill_end": cell.fill.end_color.rgb,
        "fill_type": cell.fill.fill_type,
        "horizontal": cell.alignment.horizontal,
        "vertical": cell.alignment.vertical,
    }


def test_sheet_names_parity(legacy_workbook, modular_workbook):
    """The modular export must preserve the legacy sheet order."""
    assert modular_workbook.sheetnames == legacy_workbook.sheetnames


def test_modular_workbook_uses_hydrated_repository_state(campaign_repos, modular_workbook):
    """The modular export should render live hydrated repository state."""
    kingdom = campaign_repos.kingdom.get_current_kingdom()
    food = campaign_repos.resource.get(1)

    dashboard_ws = modular_workbook["Kingdom Dashboard"]
    assert dashboard_ws["B5"].value == kingdom.treasury_silver
    assert dashboard_ws["B6"].value == kingdom.monthly_income
    assert dashboard_ws["B13"].value == kingdom.current_turn

    province_ws = modular_workbook["Provinces"]
    assert province_ws.max_row == len(campaign_repos.province.list_all()) + 1
    assert province_ws["A2"].value == campaign_repos.province.get(1).name

    resource_ws = modular_workbook["Resources"]
    assert resource_ws.max_row == len(campaign_repos.resource.list_all()) + 1
    assert resource_ws["A2"].value == food.resource_type.value
    assert resource_ws["B2"].value == food.stored

    army_ws = modular_workbook["Army Register"]
    assert army_ws.max_row == len(campaign_repos.unit.list_all()) + 1
    assert army_ws["A2"].value == campaign_repos.unit.get(1).name

    commanders_ws = modular_workbook["Commanders"]
    assert commanders_ws.max_row == len(campaign_repos.commander.list_all()) + 1

    diplomacy_ws = modular_workbook["Diplomacy & Intel"]
    assert diplomacy_ws.max_row == len(campaign_repos.relation.list_all()) + 1

    events_ws = modular_workbook["Event Log"]
    assert events_ws.max_row == 3


def test_header_style_and_width_parity(legacy_workbook, modular_workbook):
    """The modular export must preserve monolith header styling and column widths."""
    for sheet_name in legacy_workbook.sheetnames:
        legacy_ws = legacy_workbook[sheet_name]
        modular_ws = modular_workbook[sheet_name]

        for column in range(1, legacy_ws.max_column + 1):
            coordinate = legacy_ws.cell(row=1, column=column).coordinate
            legacy_cell = legacy_ws[coordinate]
            modular_cell = modular_ws[coordinate]

            assert _style_snapshot(modular_cell) == _style_snapshot(
                legacy_cell
            ), f"Style mismatch at {sheet_name}!{coordinate}"

            column_letter = legacy_cell.column_letter
            assert (
                modular_ws.column_dimensions[column_letter].width
                == legacy_ws.column_dimensions[column_letter].width
            ), f"Width mismatch at {sheet_name}!{column_letter}"


def test_legacy_header_shape_is_preserved(legacy_workbook, modular_workbook):
    """The modular export should keep the legacy workbook column counts per sheet."""
    for sheet_name in legacy_workbook.sheetnames:
        assert modular_workbook[sheet_name].max_column == legacy_workbook[sheet_name].max_column
