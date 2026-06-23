"""Phase 7 turn-simulation tests."""

import sqlite3
from pathlib import Path

from warfare_simulation.app import WarfareSimulationApp
from warfare_simulation.orchestration import GameState, SimDate
from warfare_simulation.services.campaign_service import CampaignService


ROOT = Path(__file__).resolve().parents[1]
CONFIG_DIR = ROOT / "src" / "warfare_simulation" / "config" / "data"


def test_game_state_advances_across_year_boundary():
    """The global clock should roll month 12 into year +1."""
    state = GameState(current_turn=12, current_month=12, current_year=1)

    state.advance_turn()

    assert state.current_turn == 13
    assert state.current_month == 1
    assert state.current_year == 2


def test_sim_date_formats_and_advances_as_canonical_calendar():
    """SimDate should provide the canonical observer-facing calendar behavior."""
    date = SimDate(day=28, month=2, year=4)

    next_date, month_rolled = date.advance_day()

    assert date.format() == "28/02/0004"
    assert next_date == SimDate(day=29, month=2, year=4)
    assert month_rolled is False
    assert next_date.advance_day() == (SimDate(day=1, month=3, year=4), True)


def test_game_state_advances_day_across_month_and_year_boundary():
    """Daily advancement should roll over both month and year correctly."""
    state = GameState(current_day=31, current_turn=12, current_month=12, current_year=1)

    month_rolled = state.advance_day()

    assert month_rolled is True
    assert state.current_day == 1
    assert state.current_turn == 13
    assert state.current_month == 1
    assert state.current_year == 2


def test_game_state_checkpoint_round_trip(tmp_path):
    """GameState checkpoints should save and restore the campaign clock."""
    checkpoint = tmp_path / "checkpoint.json"
    state = GameState(
        current_day=18,
        current_turn=4,
        current_month=4,
        current_year=1,
        simulation_speed="2x",
    )

    saved_path = state.save_checkpoint(checkpoint)
    restored = GameState.load_checkpoint(saved_path)

    assert restored == state


def test_phase8_simulation_speeds_match_observer_controls(tmp_path):
    """The supported speed labels should match the observer-pivot roadmap."""
    db_path = tmp_path / "phase8_speeds.db"
    app = WarfareSimulationApp(config_path=CONFIG_DIR, db_path=db_path)
    service = CampaignService(app)

    for speed, expected_interval in {
        "paused": None,
        "1x": 1000,
        "2x": 500,
        "5x": 200,
        "fast": 50,
    }.items():
        service.set_simulation_speed(speed)
        assert service.get_simulation_status().simulation_speed == speed
        assert service.get_speed_interval_ms() == expected_interval


def test_advance_turn_persists_mutable_state_to_sqlite(tmp_path):
    """Turn advancement should persist kingdom and resource changes back to SQLite."""
    db_path = tmp_path / "phase7_persisted_turn.db"
    app = WarfareSimulationApp(config_path=CONFIG_DIR, db_path=db_path)

    app.campaign.advance_turn()

    with sqlite3.connect(db_path) as conn:
        kingdom = conn.execute(
            """
            SELECT treasury_silver, current_day, current_turn, current_month, current_year
            FROM kingdom
            WHERE id = 1
            """
        ).fetchone()
        food = conn.execute(
            """
            SELECT stored
            FROM resource
            WHERE id = 1
            """
        ).fetchone()

    assert kingdom == (525700, 1, 2, 2, 1)
    assert food == (5100,)


def test_advance_day_only_triggers_monthly_economy_on_month_rollover(tmp_path):
    """Daily advancement should only apply the monthly slice when the month changes."""
    db_path = tmp_path / "phase8_daily_clock.db"
    app = WarfareSimulationApp(config_path=CONFIG_DIR, db_path=db_path)

    kingdom = app.repos.kingdom.get_current_kingdom()
    assert kingdom is not None

    for _ in range(30):
        state = app.campaign.advance_day()

    assert state.current_day == 31
    assert state.current_month == 1
    assert state.current_year == 1
    assert kingdom.treasury_silver == 520000

    state = app.campaign.advance_day()

    assert state.current_day == 1
    assert state.current_month == 2
    assert state.current_year == 1
    assert state.current_turn == 2
    assert kingdom.treasury_silver == 525700


def test_rehydrated_repositories_include_persisted_turn_state(tmp_path):
    """A restarted app should hydrate from the SQLite turn state, not the seed JSON."""
    db_path = tmp_path / "phase7_restart.db"
    app = WarfareSimulationApp(config_path=CONFIG_DIR, db_path=db_path)

    for _ in range(31):
        app.campaign.advance_day()
    restarted = WarfareSimulationApp(config_path=CONFIG_DIR, db_path=db_path)

    kingdom = restarted.repos.kingdom.get_current_kingdom()
    food = restarted.repos.resource.get(1)

    assert restarted.game_state.current_day == 1
    assert restarted.game_state.current_turn == 2
    assert restarted.game_state.current_month == 2
    assert restarted.game_state.current_year == 1
    assert kingdom.current_day == 1
    assert kingdom.current_turn == 2
    assert kingdom.current_month == 2
    assert kingdom.current_year == 1
    assert kingdom.treasury_silver == 525700
    assert food.stored == 5100


def test_advance_turn_writes_event_and_audit_logs(tmp_path):
    """Turn advancement should leave traceable event and audit records."""
    db_path = tmp_path / "phase7_audit_log.db"
    app = WarfareSimulationApp(config_path=CONFIG_DIR, db_path=db_path)

    app.campaign.advance_turn()

    with sqlite3.connect(db_path) as conn:
        events = conn.execute(
            """
            SELECT turn, category, description, impact, affected_entities
            FROM event
            ORDER BY id
            """
        ).fetchall()
        audits = conn.execute(
            """
            SELECT turn, month, year, actor, target, system, action,
                   previous_value, new_value, reason
            FROM audit_log
            ORDER BY id
            """
        ).fetchall()

    assert events[0] == (
        2,
        "Economy",
        "The Dominion of Auster collected monthly net income.",
        "Treasury changed from 520000 to 525700 silver.",
        "[1]",
    )
    assert audits[0] == (
        2,
        2,
        1,
        "kingdom:1",
        "kingdom:1.treasury_silver",
        "Economy",
        "collect_monthly_net_income",
        "520000",
        "525700",
        "Monthly income minus expenses during turn advancement.",
    )
    assert any(log[5] == "Logistics" for log in audits)


def test_advance_turn_persists_turn_summary(tmp_path):
    """Turn advancement should persist a compact summary of resolved systems."""
    db_path = tmp_path / "phase7_turn_summary.db"
    app = WarfareSimulationApp(config_path=CONFIG_DIR, db_path=db_path)

    app.campaign.advance_turn()

    with sqlite3.connect(db_path) as conn:
        summary = conn.execute(
            """
            SELECT turn, month, year, title, event_count, audit_count, highlights
            FROM turn_summary
            ORDER BY id
            """
        ).fetchone()

    assert summary[:6] == (2, 2, 1, "Turn 2 Summary", 4, 8)
    assert "The Dominion of Auster collected monthly net income." in summary[6]
    assert "Kingdom of Norland chose watch_rivals" in summary[6]


def test_rehydrated_repositories_include_turn_summaries(tmp_path):
    """A restarted app should hydrate persisted turn summaries from SQLite."""
    db_path = tmp_path / "phase7_summary_restart.db"
    app = WarfareSimulationApp(config_path=CONFIG_DIR, db_path=db_path)

    app.campaign.advance_turn()
    restarted = WarfareSimulationApp(config_path=CONFIG_DIR, db_path=db_path)

    summary = restarted.repos.turn_summary.get_latest()

    assert summary.turn == 2
    assert summary.event_count == 4
    assert summary.audit_count == 8
    assert "Logistics resolved 4 resource update(s)." in summary.narrative
    assert "FactionIntent recorded 3 observer note(s)." in summary.narrative


def test_pulse_scheduler_reports_ordered_boundaries_once():
    """Scheduler should emit ordered pulse boundaries without duplicates."""
    from warfare_simulation.orchestration import PulseScheduler

    scheduler = PulseScheduler()
    calls = []
    for pulse in ("daily", "weekly", "monthly", "seasonal", "yearly"):
        scheduler.register(pulse, lambda date, pulse=pulse: calls.append((pulse, date.format())))

    previous_date = SimDate(day=31, month=12, year=1)
    current_date = SimDate(day=1, month=1, year=2)

    report = scheduler.run_due_pulses(previous_date, current_date)
    duplicate_report = scheduler.run_due_pulses(previous_date, current_date)

    assert report.pulses == ("daily", "monthly", "seasonal", "yearly")
    assert duplicate_report.pulses == ()
    assert calls == [
        ("daily", "01/01/0002"),
        ("monthly", "01/01/0002"),
        ("seasonal", "01/01/0002"),
        ("yearly", "01/01/0002"),
    ]


def test_advance_day_records_weekly_monthly_seasonal_and_yearly_pulses(tmp_path):
    """Daily orchestration should expose the scheduler pulse report for boundaries."""
    db_path = tmp_path / "phase9_pulses.db"
    app = WarfareSimulationApp(config_path=CONFIG_DIR, db_path=db_path)

    for _ in range(6):
        app.campaign.advance_day()
    assert app.campaign.last_pulse_report.pulses == ("daily", "weekly")

    for _ in range(24):
        app.campaign.advance_day()
    assert app.campaign.last_pulse_report.pulses == ("daily",)

    app.campaign.advance_day()
    assert app.campaign.last_pulse_report.pulses == ("daily", "monthly")

    kingdom = app.repos.kingdom.get_current_kingdom()
    kingdom.current_day = 31
    kingdom.current_month = 12
    kingdom.current_year = 1
    kingdom.current_turn = 12
    app.repos.kingdom.update(kingdom)
    app.game_state.sync_from_kingdom(kingdom)
    app.campaign.advance_day()

    assert app.campaign.last_pulse_report.pulses == ("daily", "monthly", "seasonal", "yearly")
    assert app.game_state.formatted_date() == "01/01/0002"


def test_event_metadata_survives_restart_and_exports_causal_details(tmp_path):
    """Structured event metadata should persist and remain observer-readable."""
    db_path = tmp_path / "phase10_event_metadata.db"
    app = WarfareSimulationApp(config_path=CONFIG_DIR, db_path=db_path)

    app.campaign.advance_turn()
    restarted = WarfareSimulationApp(config_path=CONFIG_DIR, db_path=db_path)
    event = next(
        event
        for event in restarted.repos.event.list_all()
        if event.source_system == "Economy"
    )

    assert event.day == 1
    assert event.month == 2
    assert event.year == 1
    assert event.actor == "kingdom:1"
    assert event.target == "kingdom:1.treasury_silver"
    assert event.source_system == "Economy"
    assert event.cause_chain[:2] == ["monthly_pulse", "collect_monthly_net_income"]
    assert event.effect_summary == "Treasury changed from 520000 to 525700 silver."

    workbook = app.export_campaign(tmp_path / "metadata_export.xlsx")
    from openpyxl import load_workbook

    ws = load_workbook(workbook)["Event Log"]
    assert ws.max_column == 4
    assert "01/02/0001 | kingdom:1 → kingdom:1.treasury_silver" in ws["C2"].value
    assert "Cause: monthly_pulse → collect_monthly_net_income" in ws["C2"].value


def test_daily_monthly_pulse_writes_causal_logs_and_summary(tmp_path):
    """Monthly work reached through day ticks should be as auditable as advance_turn()."""
    db_path = tmp_path / "phase10_daily_pulse_logs.db"
    app = WarfareSimulationApp(config_path=CONFIG_DIR, db_path=db_path)

    for _ in range(31):
        app.campaign.advance_day()

    events = app.repos.event.get_by_turn(2)
    audits = app.repos.audit_log.get_by_turn(2)
    summaries = app.repos.turn_summary.get_by_turn(2)

    assert any(event.source_system == "Economy" for event in events)
    assert any(event.source_system == "Logistics" for event in events)
    assert any("daily_scheduler" in event.cause_chain for event in events)
    assert any(audit.system == "Economy" for audit in audits)
    assert any(audit.system == "Logistics" for audit in audits)
    assert summaries and summaries[-1].event_count == len(events)


def test_dashboard_event_rows_expose_observer_causality_fields(tmp_path):
    """The UI service should provide event-feed fields needed by the observatory."""
    db_path = tmp_path / "phase10_dashboard_event_feed.db"
    app = WarfareSimulationApp(config_path=CONFIG_DIR, db_path=db_path)
    service = CampaignService(app)

    for _ in range(31):
        app.campaign.advance_day()

    row = next(row for row in service.get_events() if row.source_system in {"Economy", "Logistics"})

    assert row.date == "01/02/0001"
    assert row.actor.startswith("kingdom:")
    assert row.target
    assert row.source_system in {"Economy", "Logistics"}
    assert "daily_scheduler" in row.cause
    assert row.impact


def test_phase10_dedicated_observer_logs_persist_by_stream(tmp_path):
    """Dedicated observer streams should persist and rehydrate beside generic audits."""
    db_path = tmp_path / "phase10_observer_streams.db"
    app = WarfareSimulationApp(config_path=CONFIG_DIR, db_path=db_path)

    app.campaign.advance_turn()

    economics_logs = app.repos.observer_log.get_by_stream("economics")
    assert economics_logs
    assert any(log.source_system == "Economy" for log in economics_logs)
    logistics_logs = app.repos.observer_log.get_by_stream("logistics")
    assert logistics_logs
    assert any(log.source_system == "Logistics" for log in logistics_logs)
    assert economics_logs[0].details

    restarted = WarfareSimulationApp(config_path=CONFIG_DIR, db_path=db_path)
    rehydrated = restarted.repos.observer_log.get_by_stream("economics")
    assert len(rehydrated) == len(economics_logs)
    assert len(restarted.repos.observer_log.get_by_stream("logistics")) == len(logistics_logs)
    assert rehydrated[0].summary
    assert "new_" in " ".join(rehydrated[0].details.keys())


def test_observer_summary_generator_builds_daily_weekly_monthly_views(tmp_path):
    """The observer layer should expose readable daily/weekly/monthly summaries."""
    db_path = tmp_path / "phase10_observer_summaries.db"
    app = WarfareSimulationApp(config_path=CONFIG_DIR, db_path=db_path)
    service = CampaignService(app)

    app.campaign.advance_turn()
    summaries = service.get_observer_summaries()

    assert [summary.period for summary in summaries] == ["daily", "weekly", "monthly"]
    monthly = summaries[-1]
    assert monthly.event_count == 4
    assert monthly.audit_count == 8
    assert monthly.observer_log_count == 8
    assert "Active streams: diplomacy, economics, logistics." in monthly.narrative
    assert "The Dominion of Auster collected monthly net income." in monthly.highlights
    assert "Kingdom of Norland chose watch_rivals" in monthly.highlights


def test_phase1a_autonomous_faction_intents_are_logged(tmp_path):
    """Living Chronicle Phase 1A should turn faction pressure into auditable intents."""
    db_path = tmp_path / "phase1a_faction_intents.db"
    app = WarfareSimulationApp(config_path=CONFIG_DIR, db_path=db_path)

    app.campaign.advance_turn()

    diplomacy_logs = app.repos.observer_log.get_by_stream("diplomacy")
    faction_events = [
        event for event in app.repos.event.list_all()
        if event.source_system == "FactionIntent"
    ]
    faction_audits = [
        audit for audit in app.repos.audit_log.list_all()
        if audit.system == "FactionIntent"
    ]

    assert len(diplomacy_logs) == 3
    assert len(faction_events) == 3
    assert len(faction_audits) == 3
    assert all(log.details["valid"] for log in diplomacy_logs)
    assert {log.details["intent_type"] for log in diplomacy_logs} == {"watch_rivals", "rebuild_forces"}
    assert all("evaluate_faction_pressure" in event.cause_chain for event in faction_events)

    restarted = WarfareSimulationApp(config_path=CONFIG_DIR, db_path=db_path)
    assert len(restarted.repos.observer_log.get_by_stream("diplomacy")) == 3


def test_phase1a_twelve_month_unattended_run_keeps_intents_auditable(tmp_path):
    """A one-year observer run should produce monthly faction intents without player orders."""
    db_path = tmp_path / "phase1a_twelve_month_run.db"
    app = WarfareSimulationApp(config_path=CONFIG_DIR, db_path=db_path)

    for _ in range(12):
        app.campaign.advance_turn()

    faction_events = [
        event for event in app.repos.event.list_all()
        if event.source_system == "FactionIntent"
    ]
    summaries = app.repos.turn_summary.list_all()

    assert app.game_state.current_turn == 13
    assert len(faction_events) == 36
    assert len(summaries) == 12
    assert all(event.cause_chain[-1] == "validate_intent" for event in faction_events)
    assert all(summary.event_count >= 4 for summary in summaries)

def test_event_feed_limits_recent_rows_for_long_run_readability(tmp_path):
    """Long observer runs should not force the dashboard event feed to render everything."""
    db_path = tmp_path / "phase10_event_feed_limit.db"
    app = WarfareSimulationApp(config_path=CONFIG_DIR, db_path=db_path)
    service = CampaignService(app)

    for _ in range(3):
        app.campaign.advance_turn()

    rows = service.get_events(limit=2)

    assert len(rows) == 2
    assert rows[0].turn == 4
    assert rows[1].turn == 4
    assert all(row.source_system == "FactionIntent" for row in rows)
