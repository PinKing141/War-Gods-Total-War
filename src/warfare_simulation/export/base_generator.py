"""
Abstract base class for all sheet generators.
"""

from abc import abstractmethod
from typing import Any, List, Optional

from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from warfare_simulation.core.base import SheetGenerator as CoreSheetGenerator

from .styles import StyleManager


class SheetGenerator(CoreSheetGenerator):
    """
    Base class for a domain-specific spreadsheet generator.
    Extends the core SheetGenerator with workbook and style handling.
    """

    def __init__(self, sheet_name: str, workbook: Workbook, style_manager: StyleManager):
        """
        Initialize the sheet generator.

        Args:
            sheet_name: The name of the sheet to generate.
            workbook: The openpyxl Workbook object to add the sheet to.
            style_manager: The StyleManager for consistent formatting.
        """
        super().__init__(sheet_name)
        self.wb = workbook
        self.styles = style_manager
        self.ws: Optional[Worksheet] = None  # Will be set during generation

    @abstractmethod
    def generate(self) -> None:
        """
        Generate the sheet, format it, and populate it with data.
        """
        self.ws = self.wb.create_sheet(title=self.sheet_name)
        super().generate()

    def _format_header(self, headers: List[str], column_widths: Optional[List[int]] = None) -> None:
        """Appends and formats the header row for the current worksheet."""
        if not self.ws:
            raise RuntimeError("Worksheet has not been created. Call generate() first.")

        self.ws.append(headers)

        for col_num, cell in enumerate(self.ws[1], 1):
            cell.font = self.styles.header_font()
            cell.fill = self.styles.header_fill()
            cell.alignment = self.styles.center_alignment()

            column_letter = get_column_letter(col_num)
            if column_widths and col_num <= len(column_widths):
                self.ws.column_dimensions[column_letter].width = column_widths[col_num - 1]
            else:
                self.ws.column_dimensions[column_letter].width = 18  # Default from monolith

    def _append_data(self, data: List[List[Any]]) -> None:
        """Appends data rows to the current worksheet."""
        if not self.ws:
            raise RuntimeError("Worksheet has not been created. Call generate() first.")

        for row in data:
            self.ws.append(row)