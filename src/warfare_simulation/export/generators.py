from openpyxl import Workbook
from warfare_simulation.domain.geography.repository import ProvinceRepository
from openpyxl.worksheet.worksheet import Worksheet
from warfare_simulation.export.base_generator import SheetGenerator
from warfare_simulation.domain.kingdom.repository import KingdomRepository

class DashboardGenerator(SheetGenerator):
    def __init__(self, kingdom_repo: KingdomRepository):
        self.kingdom_repo = kingdom_repo

    @property
    def sheet_name(self) -> str: return "Kingdom Dashboard"

    def generate(self, workbook: Workbook) -> Worksheet:
        ws = workbook.create_sheet(title=self.sheet_name)
        
        active_kingdom = self.kingdom_repo.get_current_kingdom()
        
        headers = ["Category", "Value", "Notes"]
        data = [
            ["Kingdom", getattr(active_kingdom, 'name', "The Dominion of Auster"), "Motto: Order Before Ambition"],
            ["Ruler", "Lord Protector Favour", "Age: 21, Stoic, Calculating"],
            ["Population", "=SUM(Provinces!B2:B5)", "Total citizens"],
            ["Treasury (Silver)", 520000, "Vast reserves of silver and iron"],
            ["Monthly Income", 18500, "Taxes & Tyra Trade"],
            ["Monthly Expenses", 12800, "Army Upkeep & Logistics"],
            ["Net Income", "=B5-B6", "Monthly surplus/deficit"],
            ["Grain Stores (Months)", 14, "Highly resilient to sieges"],
            ["Army Size", "=SUM('Army Register'!B2:B10)", "Professional Standing Army"],
            ["National Morale", "87%", "Disciplined and steady"],
            ["Noble Loyalty", "83%", "House Kael & Thorne hold sway"],
            ["Current Turn", 1, "Year 1, Month 1, Early Spring"]
        ]
        
        self._write_table(ws, headers, data, column_widths=[25, 20, 40])
        return ws

class ProvincesGenerator(SheetGenerator):
    def __init__(self, province_repo: ProvinceRepository):
        self.province_repo = province_repo

    @property
    def sheet_name(self) -> str: return "Provinces"

    def generate(self, workbook: Workbook) -> Worksheet:
        ws = workbook.create_sheet(title=self.sheet_name)
        
        headers = ["Province", "Population", "Fort Level", "Food Stored", "Monthly Tax", "Loyalty", "Garrison", "Governor"]
        data = []
        
        # 1. Fetch live data from the database!
        provinces = self.province_repo.list_all()
        
        # 2. Map domain models to Excel rows
        for prov in provinces:
            data.append([
                prov.name,
                prov.population,
                prov.fort_level,
                prov.food_stored,
                prov.monthly_tax,
                prov.loyalty,
                prov.garrison_size,
                prov.governor_name
            ])
            
        self._write_table(ws, headers, data)
        return ws

class ResourcesGenerator(SheetGenerator):
    @property
    def sheet_name(self) -> str: return "Resources"
    def generate(self, workbook: Workbook) -> Worksheet:
        return workbook.create_sheet(title=self.sheet_name)

class ArmyRegisterGenerator(SheetGenerator):
    @property
    def sheet_name(self) -> str: return "Army Register"
    def generate(self, workbook: Workbook) -> Worksheet:
        return workbook.create_sheet(title=self.sheet_name)

class CommandersGenerator(SheetGenerator):
    @property
    def sheet_name(self) -> str: return "Commanders"
    def generate(self, workbook: Workbook) -> Worksheet:
        return workbook.create_sheet(title=self.sheet_name)

class DiplomacyGenerator(SheetGenerator):
    @property
    def sheet_name(self) -> str: return "Diplomacy & Intel"
    def generate(self, workbook: Workbook) -> Worksheet:
        return workbook.create_sheet(title=self.sheet_name)

class LogisticsGenerator(SheetGenerator):
    @property
    def sheet_name(self) -> str: return "Logistics & Projects"
    def generate(self, workbook: Workbook) -> Worksheet:
        return workbook.create_sheet(title=self.sheet_name)

class EventLogGenerator(SheetGenerator):
    @property
    def sheet_name(self) -> str: return "Event Log"
    def generate(self, workbook: Workbook) -> Worksheet:
        return workbook.create_sheet(title=self.sheet_name)