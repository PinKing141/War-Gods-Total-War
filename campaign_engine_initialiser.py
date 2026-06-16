import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

def create_campaign_workbook():
    wb = openpyxl.Workbook()
    
    # Styling variables
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    
    def format_sheet(ws, headers, data):
        # Write headers
        ws.append(headers)
        for col_num, cell in enumerate(ws[1], 1):
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            # Adjust column width
            ws.column_dimensions[get_column_letter(col_num)].width = 18
            
        # Write data
        for row in data:
            ws.append(row)
            
    # --- 1. Dashboard ---
    ws_dash = wb.active
    ws_dash.title = "Kingdom Dashboard"
    dash_headers = ["Category", "Value", "Notes"]
    dash_data = [
        ["Kingdom", "The Dominion of Auster", "Motto: Order Before Ambition"],
        ["Ruler", "Lord Protector Favour", "Age: 21, Stoic, Calculating"],
        ["Population", "=SUM(Provinces!B2:B5)", "Total citizens"],
        ["Treasury (Silver)", 520000, "Vast reserves of silver and iron"],
        ["Monthly Income", 18500, "Taxes & Tyra Trade"],
        ["Monthly Expenses", 12800, "Army Upkeep & Logistics"],
        ["Net Income", "=B5-B6", "Monthly surplus/deficit"],
        ["Grain Stores (Months)", 14, "Highly resilient to sieges"],
        ["Army Size", "=SUM('Army Register'!B2:B10)", "Professional Standing Army"],
        ["National Morale", "87%", "Disciplined and steady"],
        ["Noble Loyalty", "83%", "House Kael & Thorne hold sway"],
        ["Current Turn", 1, "Year 1, Month 1, Early Spring"]
    ]
    format_sheet(ws_dash, dash_headers, dash_data)
    ws_dash.column_dimensions['A'].width = 25
    ws_dash.column_dimensions['B'].width = 20
    ws_dash.column_dimensions['C'].width = 40

    # --- 2. Provinces ---
    ws_prov = wb.create_sheet(title="Provinces")
    prov_headers = ["Province", "Population", "Fort Level", "Food Stored", "Monthly Tax", "Loyalty", "Garrison", "Governor"]
    prov_data = [
        ["Highreach (Capital)", 150000, 5, 25000, 8000, 95, 1000, "Lord Protector Favour"],
        ["Oakhaven (Agri)", 180000, 2, 45000, 4500, 90, 300, "Lord Vance"],
        ["Ironford (Mining)", 80000, 3, 12000, 4000, 85, 400, "Overseer Thorne"],
        ["Veyl (Border/Intel)", 40000, 4, 8000, 2000, 92, 800, "Spymaster Kael"]
    ]
    format_sheet(ws_prov, prov_headers, prov_data)

    # --- 3. Resources Ledger ---
    ws_res = wb.create_sheet(title="Resources")
    res_headers = ["Resource", "Stored", "Monthly Production", "Monthly Consumption", "Net Change", "Reserve Status"]
    res_data = [
        ["Food (Bushels)", 90000, 15000, 12500, "=C2-D2", "Stable"],
        ["Iron (Tons)", 12000, 250, 100, "=C3-D3", "Surplus"],
        ["Timber (Cords)", 8400, 140, 90, "=C4-D4", "Stable"],
        ["Stone (Tons)", 6300, 120, 40, "=C5-D5", "Surplus"],
        ["Wool/Cloth (Bolts)", 4800, 80, 50, "=C6-D6", "Stable"],
        ["Horses (Mounts)", 1150, 15, 5, "=C7-D7", "Stable"]
    ]
    format_sheet(ws_res, res_headers, res_data)

    # --- 4. Army Register ---
    ws_army = wb.create_sheet(title="Army Register")
    army_headers = ["Unit Name", "Soldiers", "Veterans", "Type", "Morale", "Fatigue", "Armour", "Location", "Commander"]
    army_data = [
        ["Vanguard I", 500, 420, "Heavy Spearmen", 92, 0, "Brigandine", "Highreach", "Lord Protector Favour"],
        ["Vanguard II", 500, 380, "Heavy Spearmen", 88, 0, "Brigandine", "Highreach", "Captain Aris"],
        ["Vanguard III", 500, 350, "Heavy Spearmen", 85, 0, "Brigandine", "Ironford", "General Thorne"],
        ["Vanguard IV", 500, 310, "Heavy Spearmen", 85, 0, "Brigandine", "Veyl", "Captain Dael"],
        ["Vanguard V", 500, 290, "Heavy Spearmen", 82, 0, "Brigandine", "Oakhaven", "Lieutenant Vane"],
        ["Crossbows I", 500, 250, "Ranged", 89, 0, "Gambeson", "Veyl", "Captain Sylas"],
        ["Crossbows II", 500, 220, "Ranged", 86, 0, "Gambeson", "Highreach", "Lieutenant Corren"],
        ["Iron Guard", 500, 450, "Heavy Infantry", 95, 0, "Plate/Mail", "Highreach", "Lord Protector Favour"],
        ["Thorne's Riders", 500, 300, "Medium Cavalry", 84, 0, "Mail", "Ironford", "General Thorne"]
    ]
    format_sheet(ws_army, army_headers, army_data)

    # --- 5. Commanders ---
    ws_com = wb.create_sheet(title="Commanders")
    com_headers = ["Name", "Role", "Leadership", "Tactics", "Logistics", "Loyalty", "Status", "Traits"]
    com_data = [
        ["Lord Protector Favour", "Sovereign", 95, 97, 96, 100, "Active", "Defensive Strategist, Patient, Over-analytical"],
        ["Marcus Thorne", "General (Cavalry)", 84, 79, 72, 88, "Active", "Bold, Inspiring, Prideful"],
        ["Elias Kael", "Spymaster", 75, 82, 88, 92, "Active", "Deceptive, Cautious, Calculating"],
        ["Julian Vance", "Steward", 60, 50, 94, 95, "Active", "Brilliant Logistician, Physically Frail"]
    ]
    format_sheet(ws_com, com_headers, com_data)

    # --- 6. Intelligence & Diplomacy ---
    ws_intel = wb.create_sheet(title="Diplomacy & Intel")
    intel_headers = ["Faction/Agent", "Type", "Opinion/Risk", "Status", "Notes/Mission"]
    intel_data = [
        ["Republic of Tyra", "Nation (Ally)", "+89", "Trade Pact Active", "Supplies gold and exotic goods for iron."],
        ["Empire of Volantis", "Nation (Rival)", "-85", "Cold War", "Expansionist empire to the south. Tensions high."],
        ["Raven 1 (Kael's Network)", "Spy", "12% Risk", "Active", "Scouting Volantis border troop movements."],
        ["Raven 2 (Kael's Network)", "Spy", "40% Risk", "Active", "Infiltrating Volantis military command."]
    ]
    format_sheet(ws_intel, intel_headers, intel_data)

    # --- 7. Logistics & Construction ---
    ws_log = wb.create_sheet(title="Logistics & Projects")
    log_headers = ["Project/Route", "Type", "Cost", "Progress", "Turns Left", "Notes"]
    log_data = [
        ["Southern Watchtowers", "Fortification", "12,000 Silver", "20%", 4, "Extending sightlines towards Volantis"],
        ["Veyl Supply Road", "Infrastructure", "8,000 Silver", "85%", 1, "Paving road to improve supply cart speed"],
        ["Supply Train Alpha", "Convoy", "500 Silver/mo", "Active", "N/A", "Route: Highreach -> Veyl (Food & Arrows)"]
    ]
    format_sheet(ws_log, log_headers, log_data)

    # --- 8. Event Log ---
    ws_events = wb.create_sheet(title="Event Log")
    events_headers = ["Turn", "Category", "Event Details", "Impact"]
    events_data = [
        [0, "System", "Campaign Initialised.", "The Dominion of Auster is ready."],
        [1, "Diplomacy", "Tyra trade envoy arrives in Highreach.", "+500 Silver this turn."]
    ]
    format_sheet(ws_events, events_headers, events_data)

    # Save workbook
    file_name = "Auster_Campaign_Engine.xlsx"
    wb.save(file_name)
    print(f"Spreadsheet '{file_name}' generated successfully.")

if __name__ == "__main__":
    create_campaign_workbook()